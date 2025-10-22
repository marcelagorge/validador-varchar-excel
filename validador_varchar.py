"""
Validador de limites VARCHAR em planilha Excel.

Este script verifica se algum valor em uma planilha do Excel ultrapassa o limite
definido na primeira linha (ex: 'VARCHAR(45)').
"""

from openpyxl import load_workbook
import re  # para extrair números do VARCHAR

# Caminho do arquivo Excel
arquivo = r'C:\Users\marce\Desktop\TESTEPRODUTO.xlsx'  # ajuste conforme necessário
wb = load_workbook(arquivo)
ws = wb.active

# Primeira linha contém os limites como "VARCHAR(45)"
limites = []
for cell in ws[1]:
    if cell.value:
        # Extrai o número dentro de VARCHAR(...)
        m = re.search(r'\((\d+)\)', str(cell.value))
        if m:
            limites.append(int(m.group(1)))
        else:
            limites.append(0)
    else:
        limites.append(0)

erros = []

# Percorre as linhas a partir da segunda
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    for j, valor in enumerate(row):
        if valor is not None and len(str(valor)) > limites[j]:
            erros.append({
                'Linha': i,
                'Coluna': ws.cell(row=1, column=j+1).column_letter,
                'Valor': valor,
                'Tamanho': len(str(valor)),
                'Limite': limites[j]
            })

# Mostra os erros
if erros:
    print("\n⚠️ Valores que ultrapassam os limites:")
    for e in erros:
        print(f"Linha {e['Linha']}, Coluna {e['Coluna']}: '{e['Valor']}' "
              f"(tamanho={e['Tamanho']}, limite={e['Limite']})")
else:
    print("✅ Nenhum valor ultrapassa os limites.")
